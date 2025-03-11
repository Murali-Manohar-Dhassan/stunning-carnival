import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Read the input data from the Excel file
input_file_path = "C:/Users/DELL/Frequency Allottment/slot_allocation.xlsx"  # Replace with your input Excel file path
input_df = pd.read_excel(input_file_path)

# Generate all possible slot names (P1 to P45)
all_slots = [f"P{i}" for i in range(1, 46)]

# Extract all unique station names
all_stations = input_df["Station"].unique()

# Create the output DataFrame with the slots
output_df = pd.DataFrame(index=all_slots, columns=["Slot"])
output_df["Slot"] = all_slots  # Populate the 'Slot' column

# Populate the slots for each station
for _, row in input_df.iterrows():
    station = row["Station"]
    stationary_slots = str(row["Stationary Kavach Slots"]).split(", ")
    onboard_slots = str(row["Onboard Kavach Slots"]).split(", ")
    
    # Combine Stationary and Onboard slots into one column
    for slot in stationary_slots:
        if slot in all_slots:
            output_df.loc[slot, f"{station}_Stationary_Onboard"] = slot

    for slot in onboard_slots:
        if slot in all_slots:
            # Only populate if no entry exists (to avoid overwriting stationary slots)
            if pd.isna(output_df.loc[slot, f"{station}_Stationary_Onboard"]):
                output_df.loc[slot, f"{station}_Stationary_Onboard"] = slot

# Replace NaN with empty strings for better presentation
output_df.fillna("", inplace=True)

# Save the DataFrame to an Excel file using openpyxl to apply formatting
output_file_path = "C:/Users/DELL/Frequency Allottment/output_kavach_slots_colored.xlsx"  # Output file path
output_df.to_excel(output_file_path, index=False)

# Load the saved workbook to apply color formatting
wb = openpyxl.load_workbook(output_file_path)
ws = wb.active

# Define the yellow fill color for stationary slots
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Apply the yellow color to stationary slots
for row in range(2, len(all_slots) + 2):  # Skip the header row
    for station in all_stations:
        cell = ws.cell(row=row, column=ws[1].value.split("_")[0].index(f"{station}_Stationary_Onboard")+2)
        slot = ws.cell(row=row, column=1).value
        if slot in str(input_df["Stationary Kavach Slots"]).split(", "):
            cell.fill = yellow_fill

# Save the workbook with color formatting applied
wb.save(output_file_path)

print(f"Output saved to {output_file_path}")
