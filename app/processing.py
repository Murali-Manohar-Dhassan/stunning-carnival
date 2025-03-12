import os
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill

# Define file paths
BASE_DIR = os.getcwd()
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")  # Define an upload folder
os.makedirs(UPLOAD_FOLDER, exist_ok=True)  # Ensure folder exist
INPUT_FILE = os.path.join(UPLOAD_FOLDER, "slot_allocation.xlsx")
OUTPUT_FILE = os.path.join(UPLOAD_FOLDER, "output_kavach_slots_colored.xlsx")

# Step 1: Slot Allocation
def allocate_slots(stations, max_slots=45, max_frequencies=7):
    
    allocations = []
    current_frequency = 1
    
    # Initialize allocation trackers for stationary and onboard Kavach slots
    station_alloc = [0] * max_slots  # Track allocated slots
    onboard_alloc = [0] * max_slots  # Track onboard slots

    def next_frequency():
        nonlocal current_frequency
        current_frequency += 1
        if current_frequency > max_frequencies:
            current_frequency = 1
        station_alloc[:] = [0] * max_slots
        onboard_alloc[:] = [0] * max_slots

    for station in stations:
        station_name = station["name"]
        station_slots = station["stationSlots"]
        onboard_slots = station["onboardSlots"]
        station_slot_range = []  # To hold allocated stationary slot range for each station
        onboard_slot_allocations = []  # To hold allocated onboard slots for each station

        # Check if there are enough available slots, else move to next frequency
        available_station_slots = station_alloc.count(0)
        available_onboard_slots = onboard_alloc.count(0)

        if station_slots > available_station_slots or onboard_slots > available_onboard_slots:
            next_frequency()
            available_station_slots = station_alloc.count(0)
            available_onboard_slots = onboard_alloc.count(0)
        
        # Allocate stationary Kavach slots
        allocated_station_slots = 0
        for i in range(max_slots):
            if station_alloc[i] == 0 and allocated_station_slots < station_slots:
                station_alloc[i] = station_name
                station_slot_range.append(f"P{i+1}")
                allocated_station_slots += 1

        # Allocate onboard Kavach slots dynamically
        allocated_onboard_slots = 0
        available_onboard_slot_monitoring = available_onboard_slots
        onboard_slots_monitoring = onboard_slots

        i = 0
        while allocated_onboard_slots < onboard_slots and i < max_slots:
            # Skip if the slot is already occupied by stationary allocation
            if station_alloc[i] == station_name:
                i += 1
                continue

            if available_onboard_slot_monitoring / 2 >= onboard_slots:
                # Alternate Allocation
                if onboard_alloc[i] == 0:
                    onboard_alloc[i] = station_name
                    onboard_slot_allocations.append(f"P{i+1}")
                    allocated_onboard_slots += 1
                    # Skip the next slot to maintain alternation
                    available_onboard_slot_monitoring -= 1
                    onboard_slots_monitoring -= 1
                    i += 2
                else:
                    i += 1
            else:
                # Continuous Allocation
                if onboard_alloc[i] == 0:
                    onboard_alloc[i] = station_name
                    onboard_slot_allocations.append(f"P{i+1}")
                    allocated_onboard_slots += 1
                    available_onboard_slot_monitoring -= 1
                    onboard_slots_monitoring -= 1
                i += 1

        # Append station allocation details to the report
        allocations.append({
            "Station": station_name,
            "Frequency": current_frequency,
            "Stationary Kavach Slots": ", ".join(station_slot_range),
            "Onboard Kavach Slots": ", ".join(onboard_slot_allocations)
        })
    return allocations

# Step 2: Generate Excel File
def generate_excel(stations):
    allocations = allocate_slots(stations)  # Get allocations
    print("Generating Excel file...")

    try:
        # Convert allocations into DataFrame
        df = pd.DataFrame(allocations)

        # Ensure the "Frequency" column exists
        if "Frequency" not in df.columns:
            df["Frequency"] = 1  # Default frequency if missing

        # Add "Slot" column
        df.insert(0, "Slot", [f"P{i}" for i in range(1, len(df) + 1)])

        # Save to Excel
        print(f"Saving unformatted Excel file to: {INPUT_FILE}")
        df.to_excel(INPUT_FILE, index=False)

        # Apply color scheme AFTER making sure "Frequency" exists
        apply_color_scheme()

        print(f"Excel file saved successfully: {OUTPUT_FILE}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

    return OUTPUT_FILE  # Return final colored Excel file


# Step 3: Apply Color Coding
def apply_color_scheme():
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError("Generated Excel file not found!")

    
    # Define color mapping for different frequency values
    color_map = {
        1: "FFFF00",  # Yellow
        2: "0000FF",  # Blue
        3: "FFA500",  # Orange
        4: "FF0000",  # Red
        5: "800080",  # Purple
        6: "FFC0CB",  # Pink
        7: "008000"   # Green
    }

    df = pd.read_excel(INPUT_FILE)
    
    # Generate all possible slot names (P1 to P45)
    all_slots = [f"P{i}" for i in range(1, 46)]

    # Extract all unique station names
    all_stations = df["Station"].unique()

    # Initialize the output DataFrame with all slots and station columns
    output_df = pd.DataFrame(index=all_slots, columns=["Slot"])
    output_df["Slot"] = all_slots  # Populate the 'Slot' column with P1 to P45

    # Populate the slots for each station
    for _, row in df.iterrows():
        station = row["Station"]
        frequency = row["Frequency"]
        stationary_slots = str(row["Stationary Kavach Slots"]).split(", ")
        onboard_slots = str(row["Onboard Kavach Slots"]).split(", ")
        
        # Get the color based on the frequency
        color_code = color_map.get(frequency, "FFFFFF")  # Default to white if frequency not in range
        
        # Combine Stationary and Onboard slots into one column
        for slot in stationary_slots:
            if slot in all_slots:
                output_df.loc[slot, f"{station}"] = slot

        for slot in onboard_slots:
            if slot in all_slots:
                # Only populate if no entry exists (to avoid overwriting stationary slots)
                if pd.isna(output_df.loc[slot, f"{station}"]):
                    output_df.loc[slot, f"{station}"] = slot

    # Replace NaN with empty strings for better presentation
    output_df.fillna("", inplace=True)

    # Save the DataFrame to an Excel file using openpyxl to apply formatting
    output_df.to_excel(OUTPUT_FILE, index=False)
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Apply the color formatting for stationary slots based on frequency
    for row in range(2, len(all_slots) + 2):  # Skip the header row
        for station in all_stations:
            # Check if the slot is a stationary slot for the current station
            cell_value = ws.cell(row=row, column=1).value
            if cell_value in str(df[df["Station"] == station]["Stationary Kavach Slots"].values[0]).split(", "):
                # Apply the color formatting for stationary slots
                frequency = df[df["Station"] == station]["Frequency"].values[0]
                color_code = color_map.get(frequency, "FFFFFF")
                # Get the corresponding column for the current station
                column_index = output_df.columns.get_loc(f"{station}") + 1
                cell = ws.cell(row=row, column=column_index)
                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

    # Save final colored Excel file
    wb.save(OUTPUT_FILE)

