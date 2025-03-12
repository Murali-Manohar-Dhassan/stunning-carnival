import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def apply_color_scheme():
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError("Generated Excel file not found!")

    # Define color mapping for different frequency values
    color_map = {
        1: "FFFF00",  # Yellow
        2: "0000FF",  # Blue
        3: "FFA500",  # Orange
        4: "FF0000",  # Red
        5: "800080",  # Purple
        6: "FFC0CB",  # Pink
        7: "008000"   # Green
    }

    # Load Data
    df = pd.read_excel(INPUT_FILE)
    
    # Generate all possible slot names (P1 to P45)
    all_slots = [f"P{i}" for i in range(1, 46)]

    # Create a structured DataFrame for coloring
    output_df = pd.DataFrame(index=all_slots, columns=["Slot"])
    output_df["Slot"] = all_slots  # Populate Slot column

    # Populate slots and add station columns dynamically
    station_columns = set()
    for _, row in df.iterrows():
        station = row["Station"]
        frequency = row["Frequency"]
        station_columns.add(station)

        stationary_slots = str(row["Stationary Kavach Slots"]).split(", ")
        onboard_slots = str(row["Onboard Kavach Slots"]).split(", ")
        
        # Assign slot allocations
        for slot in stationary_slots + onboard_slots:  # Merge both slot types
            if slot in all_slots:
                output_df.loc[slot, station] = slot  # Assign slot to respective station

    # Convert station column set to list
    station_columns = list(station_columns)
    output_df = output_df.reindex(columns=["Slot"] + station_columns)  # Ensure correct order

    # Save initial structured DataFrame to Excel
    output_df.to_excel(OUTPUT_FILE, index=False)

    # Load the Excel file for formatting
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Apply color formatting for allocated slots
    for row in ws.iter_rows(min_row=2, max_row=len(all_slots) + 1, min_col=2, max_col=len(station_columns) + 1):
        for cell in row:
            if cell.value:
                station = ws.cell(row=1, column=cell.column).value  # Get station name from header
                frequency = df[df["Station"] == station]["Frequency"].values[0]  # Get station frequency
                
                # Apply color based on frequency
                color_code = color_map.get(frequency, "FFFFFF")  # Default white if not found
                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

    # Save final colored Excel file
    wb.save(OUTPUT_FILE)
